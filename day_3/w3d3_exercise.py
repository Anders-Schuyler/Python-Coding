# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)


# EASY MODE

# import the appropriate modules (you have 3)
import json
import requests
from openpyxl import Workbook

character_url = "https://rickandmortyapi.com/api/character"
# set up a workbook and worksheet titled "Rick and Morty Characters"
wb = Workbook()
ws_characters = wb.active
ws_characters.title = "Rick and Morty Characters"

# # assign a variable 'data' with the returned GET request
response = requests.get(character_url)
data = response.json()

# create the appropriate headers in openpyxl for all of the keys for a single character
characters = data["results"]

header_row = 1
headers = list(characters[0].keys())
for col_num, header in enumerate(headers, start=1):
    ws_characters.cell(row=header_row, column=col_num, value=header.capitalize())

# loop through all of the 'results' of the data to populate the rows and columns for each character
for row_num, character in enumerate(characters, start=2):  # Start from row 2 because row 1 is for headers
    for col_num, (key, value) in enumerate(character.items(), start=1):
        # Write the value to the cell, converting it to a string if necessary
        ws_characters.cell(row=row_num, column=col_num, value=str(value))
# NOTE: due to the headers, the rows need to be offset by one!

# MEDIUM MODE

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"
ws_locations = wb.create_sheet(title="Rick and Morty Locations")
ws_episodes = wb.create_sheet(title="Rick and Morty Episodes")

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
location_url = "https://rickandmortyapi.com/api/location"
episode_url = "https://rickandmortyapi.com/api/episode"

# populate the new worksheets appropriately with all of the data!
def populate_worksheet(api_url, worksheet):
    response = requests.get(api_url)
    data = response.json()
    items = data["results"]
    
    # Write headers (keys of the first item)
    headers = list(items[0].keys())
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header.capitalize())
    
    # Write data
    for row_num, item in enumerate(items, start=2):  # Start at row 2 because row 1 is for headers
        for col_num, (key, value) in enumerate(item.items(), start=1):
            worksheet.cell(row=row_num, column=col_num, value=str(value))

# Populate the worksheets with data
populate_worksheet(location_url, ws_locations)
populate_worksheet(episode_url, ws_episodes)

# NOTE: don't forget your headers!

# HARD MODE
# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)
# Function to populate worksheet with data from the API with pagination
def populate_worksheet(api_url, worksheet):
    row_num = 2  # Start at row 2 because row 1 is for headers
    while api_url:
        response = requests.get(api_url)
        data = response.json()
        items = data["results"]

        # Write headers if it's the first page
        if row_num == 2:
            headers = list(items[0].keys())
            for col_num, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col_num, value=header.capitalize())

        # Write data
        for item in items:
            for col_num, (key, value) in enumerate(item.items(), start=1):
                worksheet.cell(row=row_num, column=col_num, value=str(value))
            row_num += 1

        # Get the next page URL, if available
        api_url = data["info"]["next"]

# Populate the worksheets with data
populate_worksheet(character_url, ws_characters)
populate_worksheet(location_url, ws_locations)
populate_worksheet(episode_url, ws_episodes)

# NIGHTMARE
# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url
# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)
# Function to get name from URL
def get_name_from_url(url):
    if url:  # Ensure the URL exists
        response = requests.get(url)
        data = response.json()
        # Return the name if it exists, otherwise return None
        return data.get('name', None)
    return None

# Function to replace URLs with names
def replace_urls_with_names(value):
    if isinstance(value, str) and value.startswith("https://"):
        # Single URL string
        return get_name_from_url(value)
    elif isinstance(value, list) and all(isinstance(item, str) and item.startswith("https://") for item in value):
        # List of URL strings
        return [get_name_from_url(item) for item in value]
    return value

# Function to populate worksheet with data from the API with URL replacement
def populate_worksheet(api_url, worksheet):
    row_num = 2  # Start at row 2 because row 1 is for headers
    while api_url:
        response = requests.get(api_url)
        data = response.json()
        items = data["results"]

        # Write data with URL replacement
        for item in items:
            for col_num, (key, value) in enumerate(item.items(), start=1):
                # Replace URLs with names if applicable
                value = replace_urls_with_names(value)
                worksheet.cell(row=row_num, column=col_num, value=str(value))
            row_num += 1

        # Get the next page URL, if available
        api_url = data["info"]["next"]

# Populate the worksheets with data and replace URLs
populate_worksheet(character_url, ws_characters)
populate_worksheet(location_url, ws_locations)
populate_worksheet(episode_url, ws_episodes)

wb.save(r"C:/Users/ander/OneDrive/Documents/Work/Python/VetsInTech Python Course/Week 3/spreadsheets/w3d3_exercise.xlsx")
