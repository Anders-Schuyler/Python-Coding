from openpyxl import Workbook

# Magic 8 ball!

import random

fortunes = ['It is certain',
              'It is decidedly so',
              'Yes',
              'Reply hazy try again',
              'Ask again later',
              'Concentrate',
              'My reply is no',
              'Outlook not so good',
              'Very doubtful']


def get_fortune():
    return random.choice(fortunes)

fate = get_fortune()

print(fate)

# Given the function above, it will randomly choose from the fortunes list and print one.
# Using this concept we will create our own function for a new workbook in openpyxl

# Create a list called first_names and prepopulate with 10 custom first names
first_names = [
    "Napoleon",   
    "George",     
    "Alexander",  
    "Horatio",    
    "Ulysses",    
    "Douglas",    
    "Erwin",      
    "Arthur",     
    "Robert",     
    "Dwight"      
]

# Create a list called last_names and prepopulate with 10 custom last names
last_names = [
    "Patton",     
    "Montgomery", 
    "Hannibal",   
    "Bradley",    
    "Sherman",    
    "Guderian",   
    "Nimitz",     
    "Kitchener",  
    "Slim",       
    "Zhukov"      
]
# Define a function called assign_names with a parameter 'row'
def assign_names(row):
#   - cell at row=row and column=1 assign the value to str(random.randint(111111, 999999))
    ws.cell(row=row, column=1, value=str(random.randint(111111, 999999)))
#   - cell at row=row and column=2 assign the value to a random choice of first name
    ws.cell(row=row, column=2, value=random.choice(first_names))
#   - cell at row=row and column=3 assign the value to a random choice of last name
    ws.cell(row=row, column=3, value=random.choice(last_names))

# set up appropriately for a new workbook and worksheet
wb = Workbook()
ws = wb.active

# loop through the range of 1-10 and for each number in the range
# - call/invoke the assign_names function while passing in the number as the 'row' argument
for row in range(2, 12):  # This range covers rows 2 to 11, inclusive
    assign_names(row)
# save your new workbook!
wb.save(r"./Python/VetsInTech Python Course/Week 2/spreadsheets/w2d3_exercise.xlsx")