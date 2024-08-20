# ITP Week 2 Day 2 (In-Class) Practice 1
# 
# You will continue to work on the inventory spreadsheet that you created from yesterday's exercise
# import the appropriate method from the correct module
import os
from openpyxl import load_workbook

os.makedirs(r"C:\Users\ander\OneDrive\Documents\Work\VetsInTech Python Course\Week 2\spreadsheets", exist_ok=True)

# Import the workbook that you created in yesterday's exercise from
# "./spreadsheets/inventory.xlsx"
wb_path = r"C:\Users\ander\OneDrive\Documents\Work\VetsInTech Python Course\Week 2\spreadsheets\inventory.xlsx"
wb = load_workbook(wb_path)

# verify what sheet names are available
print("Sheet names:", wb.sheetnames)

# access and store the appropriate worksheet to the variable 'ws'
ws = wb["CURRENT_MONTH_INVENTORY"]

# Print out the cell values for each row
for row in ws.iter_rows(values_only=True):
    print(row)

# Create a new column within that worksheet called order_amount
ws.cell(row=1, column=ws.max_column + 1, value='order_amount')

# save the latest changes
# wb.save("./spreadsheets/inventory.xlsx")
wb.save(wb_path)