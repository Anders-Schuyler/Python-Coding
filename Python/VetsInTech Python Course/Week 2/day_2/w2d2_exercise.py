# ITP Week 2 Day 2 Exercise

# import the appropriate method from the correct module
from openpyxl import load_workbook

# Import the workbook that you updated in today's practice from
# "./spreadsheets/inventory.xlsx"
wb_path = r".\VetsInTech Python Course\Week 2\spreadsheets\inventory.xlsx"
wb = load_workbook(wb_path)

# access and store the appropriate worksheet to the variable 'ws'
ws = wb["CURRENT_MONTH_INVENTORY"]

# Define a function called add_order_amount that takes in a single parameter called 'row'
def add_order_amount(row):
    # Retrieve the values for quantity, threshold, and max_amount
    quantity = ws.cell(row=row, column=5).value
    threshold = ws.cell(row=row, column=4).value
    max_amount = ws.cell(row=row, column=3).value

    # IF the quantity is less or equal to the threshold,
    if quantity <= threshold:
        # Then calculate the order_amount (max_amount - quantity)
        order_amount = max_amount - quantity  # Corrected indentation here
        # Assign the value to that row, column 6
        ws.cell(row=row, column=6, value=order_amount)
# TIP: create variables for quantity, threshold, max_amount that retrieves the values first for cleanliness

# Ensure the new column header is added
ws.cell(row=1, column=6, value='order_amount')

# Perform a for..in loop through the range(2, len(inventory.rows))
for row in range(2, ws.max_row + 1):
#   - call the function add_order_amount() passing in the number of the range
    add_order_amount(row)


# Save the latest changes to the workbook
wb.save(wb_path)