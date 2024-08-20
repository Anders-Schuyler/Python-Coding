#importing the OpenPyXL Library
import openpyxl
from openpyxl import Workbook

#Creating a new workbook
wb = Workbook()

#Verifying the workbook type
print(type(wb))

#Accessing the active worksheet
ws = wb.active

#Creating new worksheets
ws1 = wb.create_sheet("Rugrats")
ws2 = wb.create_sheet("Hey Arnold", 0)

#Rename an existing worksheet
ws.title = "New Title"

#Accessing the renamed worksheet using its title
same_sheet = wb['New Title']

#Listing all worksheet names in the workbook
print(wb.sheetnames)

#Looping through all worksheets to print their titles
for sheet in wb:
    print(sheet.title)