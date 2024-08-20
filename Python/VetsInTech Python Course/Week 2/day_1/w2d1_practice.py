# ITP Week 2 Day 1 (In-Class) Practice

# A1. from the appropriate library, import only the Workbook
import openpyxl
from openpyxl import Workbook
# A2. Before anything, we need a workbook to work with..
wb = Workbook()
print(type(wb))
# A3. We need to interact with a single worksheet.
ws = wb.active
# A4. assign the value of "First Name" to A1
ws['A1'] = "First Name"
#More appropriate way to do it
ws.cell(row=1, column=1, value="First Name")

# A5. assign the value of "Last Name" to B1
ws['B1'] = "Last Name"
#More appropriate way to do it
ws.cell(row=1, column=2, value="Last Name")

# STOP HERE - RETURN TO LECTURE

# B1. For all of column A, starting at row 2 until row 10, make the cell values: "Gabriel" (attempt a loop)
for row in range(2, 11):  # Range is exclusive of the end, so we use 11
    ws[f'A{row}'] = "Gabriel"


last_names = ['Rolley', 'Smith', 'Balenga', 'Issac', 'Cruise', 'Depp', 'Heard', 'Qiao', 'Biden']

# B2. Loop through a range from row 2 to 10 and assign the cell value to last names according to index in column B
# NOTE: PAY ATTENTION to the starting number of the range and how it differs from the starting index of the list
for index, row in enumerate(range(2, 11)):  # We need the index to get the last names
    ws[f'B{row}'] = last_names[index]

# B3. Save the file
wb.save(r"./Work/VetsInTech Python Course/Week 2/spreadsheets/day_1_practice.xlsx")