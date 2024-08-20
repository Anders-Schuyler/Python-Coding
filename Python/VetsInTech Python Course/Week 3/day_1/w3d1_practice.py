# ITP Week 3 Day 1 Practice

# import your required modules/methods
import openpyxl
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

# Data
clefairy = {
    "id": 35,
    "name": "clefairy",
    "base_experience": 113,
    "height": 6,
    "order": 56,
    "weight": 75,
}

weedle = {
    "id": 13,
    "name": "weedle",
    "base_experience": 39,
    "height": 3,
    "order": 17,
    "weight": 32
}

# Capitalize the 'name' field in both dictionaries
clefairy["name"] = clefairy["name"].capitalize()
weedle["name"] = weedle["name"].capitalize()

# given the following items, using the methods we covered, write to openpyxl

# Write the headers to the first row
headers = list(clefairy.keys())
for col_num, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_num, value=header.capitalize())

row_count = 2

# use an external counter with just a for loop (no function)
column_count = 1

for key in clefairy:
    ws.cell(row=row_count, column=column_count, value=clefairy[key])
    column_count+=1
row_count+= 1



# create a function that takes in a pokemon
def dict_to_row(pokemon):
    global row_count
    for col_num, value in enumerate(pokemon.values(), start=1):
        ws.cell(row=row_count, column=col_num, value=value)
    row_count += 1

# call the function with weedle!
dict_to_row(weedle)

wb.save(r"./Python/VetsInTech Python Course/Week 3/spreadsheets/w3d1practice.xlsx")